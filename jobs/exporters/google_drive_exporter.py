import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from googleapiclient.http import MediaFileUpload
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from sqlalchemy import text

from config.vars import ROOT_DIR, DATA_DIR, SCRIPTS_DIR


class GoogleDriveExporter:
    def __init__(self, db_engine, google_drive_credentials, export_dir="exports"):
        """
        :param db_connection: Obiekt połączenia do bazy danych.
        :param google_drive_credentials: Ścieżka do pliku z danymi autoryzacyjnymi do Google Drive.
        :param export_dir: Katalog do zapisu plików lokalnych.
        """
        self.db_engine = db_engine
        self.credentials = Credentials.from_service_account_file(
            google_drive_credentials
        )
        self.export_dir = f"{ROOT_DIR}/{DATA_DIR}"
        os.makedirs(export_dir, exist_ok=True)

    def export_to_csv(self, df, filename):
        """Eksportuje DataFrame do pliku CSV."""
        filepath = os.path.join(self.export_dir, filename)
        df.to_csv(filepath, index=False)
        return filepath

    def run_query_from_file(self, sql_file, params=None):
        """
        Executes a SQL query from a .sql file.

        :param sql_file: Path to the .sql file.
        :param params: Optional dictionary of query parameters.
        :return: Pandas DataFrame containing the query result.
        """
        file_path = f"{ROOT_DIR}/{SCRIPTS_DIR}/{sql_file}"
        if not os.path.isfile(file_path) or not sql_file.endswith(".sql"):
            raise ValueError("Invalid SQL file path provided.")

        with open(file_path, "r") as file:
            query = file.read()

        with self.db_engine.connect() as connection:
            return pd.read_sql_query(text(query), con=connection, params=params)

    def export_to_xlsx(self, dataframes, sheet_names, filename):
        """
        Eksportuje wiele DataFrame'ów do jednego pliku XLSX z prostym formatowaniem.

        :param dataframes: Lista DataFrame'ów.
        :param sheet_names: Lista nazw arkuszy.
        :param filename: Nazwa pliku wynikowego.
        """
        filepath = os.path.join(self.export_dir, filename)
        wb = Workbook()
        for df, sheet_name in zip(dataframes, sheet_names):
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                ws.append(row)
                if r_idx == 1:  # Formatowanie nagłówka
                    for cell in ws[r_idx]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(filepath)
        return filepath

    def upload_to_google_drive(self, file_name, folder_id):
        service = build("drive", "v3", credentials=self.credentials)
        file_metadata = {"name": os.path.basename(f"{self.export_dir}/{file_name}")}
        if folder_id:
            file_metadata["parents"] = [folder_id]
        media = MediaFileUpload(f"{self.export_dir}/{file_name}", resumable=True)
        uploaded_file = (
            service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )
        return uploaded_file.get("id")
