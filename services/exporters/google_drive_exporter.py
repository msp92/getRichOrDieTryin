import logging
import datetime as dt
from typing import List, Dict
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from googleapiclient.http import MediaFileUpload
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from sqlalchemy import text

from config.vars import (
    ROOT_DIR,
    DATA_DIR,
    SCRIPTS_DIR,
    GOOGLE_DRIVE_GRODT_FOLDER_ID,
)
from models.analytics.breaks import BreaksTeamStats
from models.base import BaseMixin
from services.db import Db

EXPORT_CONFIG = {
    "breaks": {
        "queries": {
            "breaks_with_coaches": "breaks_with_coaches.sql",
            "team_stats": "BreaksTeamStats.get_all",
            "upcoming_games": "upcoming_games_with_coaches.sql",
        }
    },
    "detailed_fixtures": {
        "queries": {"detailed_fixtures": "fixtures_events_summary_new.sql"}
    },
}


class GoogleDriveExporter:
    def __init__(
        self,
        db_engine,
        export_config: Dict[str, Dict[str, Dict[str, str]]],
        gdrive_folder_id: str,
        export_dir: str = "exports",
        credentials_path: str = "google_api_credentials.json",
    ):
        """
        Initialize the Google Drive exporter with Service Account credentials.

        :param db_engine: Database engine object.
        :param export_config: Export configuration (queries for each entity).
        :param gdrive_folder_id: Google Drive folder ID for uploads.
        :param export_dir: Local directory for exported files (default: 'exports').
        :param credentials_path: Path to Google API Service Account credentials JSON file (default: 'google_api_credentials.json').
        """
        self.db_engine = db_engine
        self.export_config = export_config
        self.gdrive_folder_id = gdrive_folder_id
        self.export_dir = os.path.join(ROOT_DIR, DATA_DIR, export_dir)
        self.credentials = self.__load_credentials(credentials_path)
        self.service = build("drive", "v3", credentials=self.credentials)
        os.makedirs(self.export_dir, exist_ok=True)

    @staticmethod
    def __load_credentials(credentials_path: str) -> Credentials:
        """Load Google API Service Account credentials from a JSON file.

        :param credentials_path: Path to the credentials JSON file relative to ROOT_DIR.
        :return: Google API Credentials object.
        :raises FileNotFoundError: If the credentials file is not found.
        """
        creds_path = os.path.join(ROOT_DIR, credentials_path)
        if not os.path.exists(creds_path):
            raise FileNotFoundError(
                f"Google credentials file not found at {creds_path}"
            )
        return Credentials.from_service_account_file(
            creds_path, scopes=["https://www.googleapis.com/auth/drive"]
        )

    def run_query_from_file(self, sql_file: str, params: dict = None) -> pd.DataFrame:
        """Wykonuje zapytanie SQL z pliku."""
        file_path = os.path.join(ROOT_DIR, SCRIPTS_DIR, sql_file)
        if not os.path.isfile(file_path) or not sql_file.endswith(".sql"):
            raise ValueError(f"Invalid SQL file path: {file_path}")

        with open(file_path, "r") as file:
            query = file.read()

        with self.db_engine.connect() as connection:
            return pd.read_sql_query(text(query), con=connection, params=params)

    def export_to_csv(self, df: pd.DataFrame, filename: str) -> str:
        """Eksportuje DataFrame do pliku CSV."""
        filepath = os.path.join(self.export_dir, filename)
        df.to_csv(filepath, index=False)
        return filepath

    def export_to_xlsx(
        self, dataframes: List[pd.DataFrame], sheet_names: List[str], filename: str
    ) -> str:
        """Eksportuje wiele DataFrame’ów do jednego pliku XLSX z formatowaniem."""
        filepath = os.path.join(self.export_dir, filename)
        wb = Workbook()
        for df, sheet_name in zip(dataframes, sheet_names):
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1
            ):
                ws.append(row)
                if r_idx == 1:  # Formatowanie nagłówka
                    for cell in ws[r_idx]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(filepath)
        return filepath

    def upload_to_google_drive(self, file_name: str) -> str:
        file_metadata = {"name": os.path.basename(file_name)}
        if self.gdrive_folder_id:
            file_metadata["parents"] = [self.gdrive_folder_id]
        media = MediaFileUpload(file_name, resumable=True)
        uploaded_file = (
            self.service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )
        return uploaded_file.get("id")

    def export_all(self) -> None:
        """Eksportuje dane z konfiguracji do dwóch osobnych plików XLSX i uploaduje na Google Drive."""
        all_data = {}
        for entity_name, config in self.export_config.items():
            logging.info(f"Fetching data for {entity_name}")
            try:
                for query_name, query_source in config["queries"].items():
                    if query_source == "BreaksTeamStats.get_all":
                        all_data[query_name] = BreaksTeamStats.get_all()
                    else:
                        all_data[query_name] = self.run_query_from_file(query_source)
            except Exception as e:
                logging.error(f"Error fetching data for {entity_name}: {e}")

        if not all_data:
            logging.warning("No data to export")
            return

        timestamp = dt.datetime.now().strftime("%Y%m%d%H%M")

        # Eksport dla "breaks"
        breaks_data = {
            k: v
            for k, v in all_data.items()
            if k in ["breaks_with_coaches", "team_stats", "upcoming_games"]
        }
        if breaks_data:
            try:
                file_name = f"breaks_{timestamp}.xlsx"
                df_list = list(breaks_data.values())
                sheet_names = list(breaks_data.keys())
                filepath = self.export_to_xlsx(df_list, sheet_names, file_name)
                self.upload_to_google_drive(filepath)
                logging.info(f"Exported and uploaded {file_name} for breaks data")
            except Exception as e:
                logging.error(f"Error exporting breaks data: {e}")

        # Eksport dla "detailed_fixtures"
        fixtures_data = {k: v for k, v in all_data.items() if k == "detailed_fixtures"}
        if fixtures_data:
            try:
                file_name = f"detailed_fixtures_{timestamp}.xlsx"
                df_list = list(fixtures_data.values())
                sheet_names = list(fixtures_data.keys())
                filepath = self.export_to_xlsx(df_list, sheet_names, file_name)
                self.upload_to_google_drive(filepath)
                logging.info(
                    f"Exported and uploaded {file_name} for detailed fixtures data"
                )
            except Exception as e:
                logging.error(f"Error exporting detailed fixtures data: {e}")


if __name__ == "__main__":
    db = Db()
    BaseMixin.set_db(db)
    exporter = GoogleDriveExporter(
        db.engine, GOOGLE_API_CREDENTIALS, EXPORT_CONFIG, GOOGLE_DRIVE_GRODT_FOLDER_ID
    )
    exporter.export_all()
